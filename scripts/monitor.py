#!/usr/bin/env python3
"""
Sanctions List Monitor - MSWiA Gov.pl
Fetches the sanctions XLSX, compares with previous version,
generates changelog, and sends email notifications.
"""

import hashlib
import json
import os
import re
import smtplib
import sys
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# --- Configuration ---
BASE_URL = "https://www.gov.pl/web/mswia/lista-osob-i-podmiotow-objetych-sankcjami"
DATA_DIR = Path(__file__).parent.parent / "data"
CURRENT_FILE = DATA_DIR / "current.xlsx"
CURRENT_JSON = DATA_DIR / "current.json"
CHANGELOG_FILE = DATA_DIR / "changelog.json"
META_FILE = DATA_DIR / "meta.json"

# Email config from env
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
EMAIL_TO = os.environ.get("EMAIL_TO", "")
EMAIL_FROM = os.environ.get("EMAIL_FROM", SMTP_USER)


def fetch_xlsx_url() -> str | None:
    """Find the current XLSX download link on the MSWiA page."""
    headers = {
        "User-Agent": "SanctionsMonitor/1.0 (compliance monitoring tool)"
    }
    resp = requests.get(BASE_URL, headers=headers, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Look for xlsx link - MSWiA uses attachment links
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if ".xlsx" in href.lower():
            if href.startswith("/"):
                href = "https://www.gov.pl" + href
            elif not href.startswith("http"):
                href = "https://www.gov.pl/" + href
            return href

    # Also check for direct file references in the page
    for link in soup.find_all("a"):
        text = link.get_text(strip=True).lower()
        if "tabela" in text and "sankcyj" in text:
            href = link.get("href", "")
            if href:
                if href.startswith("/"):
                    href = "https://www.gov.pl" + href
                return href

    return None


def download_xlsx(url: str) -> bytes:
    """Download the XLSX file."""
    headers = {
        "User-Agent": "SanctionsMonitor/1.0 (compliance monitoring tool)"
    }
    resp = requests.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    return resp.content


def file_hash(data: bytes) -> str:
    """SHA256 hash of file content."""
    return hashlib.sha256(data).hexdigest()


def parse_xlsx(filepath: Path) -> list[dict]:
    """
    Parse the sanctions XLSX into a list of dicts.
    Adapts to actual column structure found in the file.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row (first row with multiple non-empty cells)
    header_idx = 0
    for i, row in enumerate(rows):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break

    # Clean headers
    raw_headers = rows[header_idx]
    headers = []
    for h in raw_headers:
        if h is None:
            headers.append("")
        else:
            # Normalize header text
            h_str = str(h).strip().lower()
            h_str = re.sub(r"\s+", " ", h_str)
            headers.append(h_str)

    # Parse data rows
    entries = []
    for row in rows[header_idx + 1:]:
        # Skip empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        entry = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                key = headers[j]
                entry[key] = str(val).strip() if val is not None else ""

        # Create a unique key from available identifying fields
        # Try common column patterns
        id_parts = []
        for key in sorted(entry.keys()):
            if any(term in key for term in ["lp", "nr", "numer"]):
                id_parts.insert(0, entry[key])
            elif any(term in key for term in ["nazwa", "imię", "imie", "nazwisko", "name"]):
                id_parts.append(entry[key])

        if not id_parts:
            # Fallback: use first 3 non-empty values
            id_parts = [v for v in entry.values() if v][:3]

        entry["_id"] = "|".join(id_parts).strip("|")

        if entry["_id"]:  # Only add entries with some identifying info
            entries.append(entry)

    wb.close()
    return entries


def compute_diff(old_entries: list[dict], new_entries: list[dict]) -> dict:
    """
    Compare old and new entries, return diff.
    Returns dict with 'added', 'removed', 'modified' lists.
    """
    old_map = {e["_id"]: e for e in old_entries if e.get("_id")}
    new_map = {e["_id"]: e for e in new_entries if e.get("_id")}

    old_ids = set(old_map.keys())
    new_ids = set(new_map.keys())

    added = []
    for _id in sorted(new_ids - old_ids):
        entry = {k: v for k, v in new_map[_id].items() if k != "_id"}
        entry["_id"] = _id
        added.append(entry)

    removed = []
    for _id in sorted(old_ids - new_ids):
        entry = {k: v for k, v in old_map[_id].items() if k != "_id"}
        entry["_id"] = _id
        removed.append(entry)

    modified = []
    for _id in sorted(old_ids & new_ids):
        old_e = {k: v for k, v in old_map[_id].items() if k != "_id"}
        new_e = {k: v for k, v in new_map[_id].items() if k != "_id"}
        if old_e != new_e:
            changes = {}
            all_keys = set(old_e.keys()) | set(new_e.keys())
            for key in all_keys:
                old_val = old_e.get(key, "")
                new_val = new_e.get(key, "")
                if old_val != new_val:
                    changes[key] = {"old": old_val, "new": new_val}
            modified.append({"_id": _id, "changes": changes})

    return {
        "added": added,
        "removed": removed,
        "modified": modified,
    }


def update_changelog(diff: dict, timestamp: str) -> None:
    """Append diff entry to changelog.json."""
    changelog = []
    if CHANGELOG_FILE.exists():
        try:
            changelog = json.loads(CHANGELOG_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, FileNotFoundError):
            changelog = []

    entry = {
        "timestamp": timestamp,
        "added_count": len(diff["added"]),
        "removed_count": len(diff["removed"]),
        "modified_count": len(diff["modified"]),
        "added": diff["added"],
        "removed": diff["removed"],
        "modified": diff["modified"],
    }

    changelog.insert(0, entry)

    # Keep last 500 changelog entries
    changelog = changelog[:500]

    CHANGELOG_FILE.write_text(
        json.dumps(changelog, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def send_email_notification(diff: dict, timestamp: str) -> None:
    """Send email notification about changes."""
    if not all([SMTP_USER, SMTP_PASS, EMAIL_TO]):
        print("Email not configured, skipping notification.")
        return

    subject = f"🔔 Zmiana na liście sankcyjnej MSWiA - {timestamp[:10]}"

    # Build HTML email
    html_parts = [
        "<html><body>",
        f"<h2>Wykryto zmiany na liście sankcyjnej MSWiA</h2>",
        f"<p><strong>Data sprawdzenia:</strong> {timestamp}</p>",
        f"<p><strong>Dodano:</strong> {len(diff['added'])} | "
        f"<strong>Usunięto:</strong> {len(diff['removed'])} | "
        f"<strong>Zmodyfikowano:</strong> {len(diff['modified'])}</p>",
        "<hr>",
    ]

    if diff["added"]:
        html_parts.append("<h3 style='color: green;'>➕ Dodane wpisy:</h3>")
        html_parts.append("<ul>")
        for entry in diff["added"]:
            display = {k: v for k, v in entry.items() if k != "_id" and v}
            html_parts.append(f"<li><strong>{entry.get('_id', 'N/A')}</strong><br>")
            for k, v in display.items():
                html_parts.append(f"&nbsp;&nbsp;{k}: {v}<br>")
            html_parts.append("</li>")
        html_parts.append("</ul>")

    if diff["removed"]:
        html_parts.append("<h3 style='color: red;'>➖ Usunięte wpisy:</h3>")
        html_parts.append("<ul>")
        for entry in diff["removed"]:
            display = {k: v for k, v in entry.items() if k != "_id" and v}
            html_parts.append(f"<li><strong>{entry.get('_id', 'N/A')}</strong><br>")
            for k, v in display.items():
                html_parts.append(f"&nbsp;&nbsp;{k}: {v}<br>")
            html_parts.append("</li>")
        html_parts.append("</ul>")

    if diff["modified"]:
        html_parts.append("<h3 style='color: orange;'>✏️ Zmodyfikowane wpisy:</h3>")
        html_parts.append("<ul>")
        for entry in diff["modified"]:
            html_parts.append(f"<li><strong>{entry['_id']}</strong><br>")
            for field, change in entry["changes"].items():
                html_parts.append(
                    f"&nbsp;&nbsp;{field}: "
                    f"<span style='color:red;text-decoration:line-through'>{change['old']}</span> → "
                    f"<span style='color:green'>{change['new']}</span><br>"
                )
            html_parts.append("</li>")
        html_parts.append("</ul>")

    html_parts.append(
        f"<hr><p><small>Źródło: <a href='{BASE_URL}'>{BASE_URL}</a></small></p>"
    )
    html_parts.append("</body></html>")

    html_body = "\n".join(html_parts)

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
        print(f"Email notification sent to {EMAIL_TO}")
    except Exception as e:
        print(f"Failed to send email: {e}")


def main():
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    print(f"[{now}] Starting sanctions list check...")

    # Ensure data dir exists
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # 1. Find XLSX URL
    print("Fetching page to find XLSX link...")
    xlsx_url = fetch_xlsx_url()
    if not xlsx_url:
        print("ERROR: Could not find XLSX link on the page!")
        sys.exit(1)
    print(f"Found XLSX URL: {xlsx_url}")

    # 2. Download file
    print("Downloading XLSX...")
    xlsx_data = download_xlsx(xlsx_url)
    new_hash = file_hash(xlsx_data)
    print(f"Downloaded {len(xlsx_data)} bytes, hash: {new_hash[:16]}...")

    # 3. Check if file changed (quick hash check)
    meta = {}
    if META_FILE.exists():
        try:
            meta = json.loads(META_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, FileNotFoundError):
            meta = {}

    old_hash = meta.get("last_hash", "")

    if new_hash == old_hash:
        print("No changes detected (same file hash). Done.")
        # Update last_checked
        meta["last_checked"] = now
        META_FILE.write_text(json.dumps(meta, indent=2), encoding="utf-8")
        return

    print("File hash changed! Analyzing differences...")

    # 4. Parse old data
    old_entries = []
    if CURRENT_JSON.exists():
        try:
            old_entries = json.loads(CURRENT_JSON.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, FileNotFoundError):
            old_entries = []

    # 5. Save new XLSX and parse
    CURRENT_FILE.write_bytes(xlsx_data)
    new_entries = parse_xlsx(CURRENT_FILE)
    print(f"Parsed {len(new_entries)} entries from new file.")

    # 6. Compute diff
    if old_entries:
        diff = compute_diff(old_entries, new_entries)
        total_changes = len(diff["added"]) + len(diff["removed"]) + len(diff["modified"])
        print(
            f"Changes: +{len(diff['added'])} added, "
            f"-{len(diff['removed'])} removed, "
            f"~{len(diff['modified'])} modified"
        )

        if total_changes > 0:
            # 7. Update changelog
            update_changelog(diff, now)
            print("Changelog updated.")

            # 8. Send email
            send_email_notification(diff, now)
        else:
            print("File hash changed but no structural changes detected.")
    else:
        print("First run - no previous data to compare. Initializing.")
        # Create initial changelog entry
        initial_diff = {"added": new_entries, "removed": [], "modified": []}
        update_changelog(initial_diff, now)

    # 9. Save current state
    CURRENT_JSON.write_text(
        json.dumps(new_entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # 10. Update meta
    meta = {
        "last_hash": new_hash,
        "last_checked": now,
        "last_changed": now,
        "xlsx_url": xlsx_url,
        "entry_count": len(new_entries),
    }
    META_FILE.write_text(json.dumps(meta, indent=2), encoding="utf-8")

    print(f"[{now}] Done.")


if __name__ == "__main__":
    main()
